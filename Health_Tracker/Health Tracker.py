from tkinter import *
from tkinter import messagebox
from tkinter import ttk 
import tkinter as tk 
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = tk.Tk()
root.geometry('1000x700')
root.title('Menu')
root.resizable(width=False,height=False)
root.config(bg='white')
# root.attributes('-alpha',0.5)

passwordwb = Workbook()
passwordwb = load_workbook('Passwords.xlsx')
passwordws = passwordwb.active # Database for student sign up (username/passwords)


database_wb = Workbook()
database_wb = load_workbook('Database.xlsx')
database_ws = database_wb.active # Database for student information 


img = PhotoImage(file = 'Amb.png')
bg_1 = PhotoImage(file = 'bg_1.png')
bg_2 = PhotoImage(file = 'bg_2.png')
logo = PhotoImage(file = 'Logo.png')


def main_window():
    header = Frame(root, width=1000, height=60, bg='#000000')
    header.grid_propagate(FALSE)
    header.grid(row=0)
    
    Label(header,text=" ", borderwidth=0,bg='#000000', fg='#FFFBF5', font=('Arial Rounded MT Bold',20)).grid(row=0, column=0,ipadx=110,ipady=17)
    main_button = Button(header, text='   MAIN MENU   ', borderwidth=0,bg='#FFFBF5', fg='#000000', font=('Arial Rounded MT Bold',20),command=lambda:main_btn())
    login_button = Button(header, text='   LOG IN   ', borderwidth=0,bg='#000000', fg='#FFFBF5', font=('Arial Rounded MT Bold',20),command=lambda:login_btn())
    signin_button = Button(header, text='   SIGN UP   ', borderwidth=0,bg='#000000', fg='#FFFBF5', font=('Arial Rounded MT Bold',20),command=lambda:signin_btn())
    main_button.grid(row=0,column=1,)
    login_button.grid(row=0,column=2)
    signin_button.grid(row=0,column=3)



    body = Frame(root, width=1000, height=600,bg='#FFFBF5')
    body.grid_propagate(False)
    body.grid(row=1)

    #####################################################

    main_menu_body = Frame(body, width=1000, height=600,bg='#FFFBF5')
    main_menu_body.grid_propagate(False)
    Label(main_menu_body, text=" Welcome to Ambasing Fitness Institute ", font=('Arial Rounded MT Bold',25),bg='#FFFBF5').grid(ipady=15)
    Label(main_menu_body, image=img, borderwidth=0,bg='white').grid(ipadx=285)
    Label(main_menu_body, font=('Arial Rounded MT Bold',8),bg='#FFFBF5', text='\n\nAmbasing Fitness Institute led by Coach Perrel Brown tracks down your health information such as nutrition and physique in a monthly basis.\nCoach Perrel Brown will help you improve your health via feedback in your contact number. ',anchor=N).grid(ipadx=10)

    main_menu_body.grid(row=0)
 
      

    #####################################################

    main_body_frame = Frame(body, width=1000, height=600,bg='orange' )
    main_menu_body.grid_propagate(False)

    Label(main_body_frame, image=bg_2).grid(row=0)

    coach = Frame(main_body_frame, width=400, height=450,bg='#525252')
    coach.grid_propagate(False)
    coach.grid(row=0,column=0,padx=50,pady=75,sticky=NW)

    # student = Frame(main_body_frame, width=400, height=450,bg='blue')
    # student.grid_propagate(False)
    # student.grid(row=0,column=1,padx=50,pady=75,sticky=NW)

    Label(coach, text='      ', font=('Rockwell',10,), bg='#525252',fg='#FFFBF5').grid(row=0,column=0, columnspan=2,sticky=W,)
    Label(coach, text='     Admin Login', font=('Rockwell',35), bg='#525252',fg='#FFFBF5').grid(row=1,column=0, columnspan=2,sticky=W,)
    Label(coach, text='\nUsername:   ', font=('Arial Rounded MT Bold',17), bg='#525252',fg='#FFFBF5').grid(row=2, column=0,sticky=W, padx=30)
    coach_username = Entry(coach, width=25, font=('Arial Rounded MT Bold',15))
    global coachname 
    coachname = coach_username.get()
    coach_username.grid(row=3, column=0,columnspan=2, sticky=E,padx=37,pady=10)
    Label(coach, text='\nPassword:   ', font=('Arial Rounded MT Bold',17), bg='#525252',fg='#FFFBF5').grid(row=4, column=0,sticky=W, padx=30)
    coach_password = Entry(coach, width=25, font=('Arial Rounded MT Bold',15), show='*')
    coach_password.grid(row=5, column=0,columnspan=2, sticky=E,padx=37,pady=10)
    Label(coach, text=' ', font=('Arial Rounded MT Bold',30), bg='#525252',fg='#FFFBF5').grid(row=6, column=0,sticky=W, padx=30)





    Button(coach, text='Login', font=('Rockwell',17), command=lambda:login(), borderwidth=0, bg='#FFFBF5').grid(row=7, column=1, sticky=N, ipadx=10, pady=15, ipady=2)

    # LOGIN COACH:      USER: "dreamy"     PASS: "000"
    def login():
        
        if coach_password.get() == '000' and coach_username.get() == 'dreamy':
            # messagebox.showinfo('Success', 'Log in successful. \nWelcome Coach!')
            coach_password.delete(0,END)
            coach_username.delete(0,END)


            root.withdraw()
            coach_menu()
            pass

        else: 
            messagebox.showerror('Error', 'Authentication failed.')


    
    # coach.grid_forget()

    #####################################################



    student = Frame(main_body_frame, width=400, height=450,bg='#525252')
    student.grid_propagate(False)
    student.grid(row=0,column=1,padx=50,pady=75,sticky=NW)

    Label(student, text='     ', font=('Rockwell',10,), bg='#525252',fg='#FFFBF5').grid(row=0,column=0, columnspan=2,sticky=W,)
    Label(student, text='     User Login', font=('Rockwell',35), bg='#525252',fg='#FFFBF5').grid(row=1,column=0, columnspan=2,sticky=W,)
    Label(student, text='\nUsername:   ', font=('Arial Rounded MT Bold',17), bg='#525252',fg='#FFFBF5').grid(row=2, column=0,sticky=W, padx=30)


    student_username = Entry(student, width=25, font=('Arial Rounded MT Bold',15))
    student_username.grid(row=3, column=0,columnspan=2, sticky=E,padx=37,pady=10)
    Label(student, text='\nPassword:   ', font=('Arial Rounded MT Bold',17), bg='#525252',fg='#FFFBF5').grid(row=4, column=0,sticky=W, padx=30)
    student_password = Entry(student, width=25, font=('Arial Rounded MT Bold',15), show='*')
    student_password.grid(row=5, column=0,columnspan=2, sticky=E,padx=37,pady=10)
    Label(student, text=' ', font=('Arial Rounded MT Bold',30), bg='#525252',fg='#FFFBF5').grid(row=6, column=0,sticky=W, padx=30)

    Button(student, text='Login', font=('Rockwell',17), command=lambda:stud_login(), borderwidth=0, bg='#FFFBF5').grid(row=7, column=1, sticky=N, ipadx=10, pady=15, ipady=2)


    # root.withdraw()



    # STUDENT LOGIN    
    def stud_login():
        if student_username.get() == '' or student_password.get() == '':
            messagebox.showerror('Error', ' Please fill up all entries.')
        else:
            Found = False
            for each_cell in range(3, passwordws.max_row+1):
               

                global global_username
                global_username = passwordws['A'+str(each_cell)].value
                b = passwordws['B'+str(each_cell)].value
                print(f'row {each_cell} =  {global_username} -> {b}  ')

                if student_username.get() == global_username and student_password.get() == b:
                    
                    Found = True
                    break
                else: 
                    Found = False
                    continue
            if Found == False:
                messagebox.showerror('Error','Login Failed: Wrong Credentials.')
                student_password.delete(0,END)

            if Found == True:
                # messagebox.showinfo('Success',f'Successfully logged in.\n Welcome! {global_username}')
                student_username.delete(0,END)
                student_password.delete(0,END)

                root.withdraw()
                student_choice()
            
    # student.grid_forget()
    main_body_frame.grid(row=0)
    main_body_frame.grid_forget()



    #####################################################

    register_window = Frame(body, width=1000, height=600,bg='#FFFBF5')
    register_window.grid_propagate(False)

    Label(register_window, image=bg_1).grid(row=0)

    rw_content = Frame(register_window, width=400, height=450, bg='#FD4556')
    rw_content.grid_propagate(False)
    rw_content.grid(row=0,padx=50,pady=75,sticky=NW)

    

    Label(rw_content, text='     ', font=('Rockwell',10,), bg='#FD4556',fg='#FFFBF5').grid(row=0,column=0, columnspan=2,sticky=W,)
    Label(rw_content, text='     Sign Up ', font=('Rockwell',35), bg='#FD4556',fg='#FFFBF5').grid(row=1,column=0, columnspan=2,sticky=W,)
    Label(rw_content, text='\nUsername:   ', font=('Arial Rounded MT Bold',17), bg='#FD4556',fg='#FFFBF5').grid(row=2, column=0,sticky=W, padx=30)
    student_username_signup = Entry(rw_content, width=25, font=('Arial Rounded MT Bold',15))
    student_username_signup.grid(row=3, column=0,columnspan=2, sticky=E,padx=37,pady=10)
    Label(rw_content, text='\nPassword:   ', font=('Arial Rounded MT Bold',17), bg='#FD4556',fg='#FFFBF5').grid(row=4, column=0,sticky=W, padx=30)
    student_password_signup = Entry(rw_content, width=25, font=('Arial Rounded MT Bold',15), show='*')
    student_password_signup.grid(row=5, column=0,columnspan=2, sticky=E,padx=37,pady=10)
    Label(rw_content, text=' ', font=('Arial Rounded MT Bold',30), bg='#FD4556',fg='#FFFBF5').grid(row=6, column=0,sticky=W, padx=30)



    Button(rw_content, text='Sign-up', font=('Rockwell',17), command=lambda:signup(), borderwidth=0, bg='#FFFBF5').grid(row=7, column=1, sticky=N, ipadx=10, pady=15, ipady=2)


    # STUDENT SIGNUP
    def signup():
        a = student_username_signup.get()
        b = student_password_signup.get()
        c = int(passwordws.cell(row=passwordws.max_row, column=3).value) + 1

        if student_username_signup.get() == '':
            messagebox.showerror('Error', 'Please Enter Your Username')
        elif student_password_signup.get() == '':
            messagebox.showerror('Error', 'Please Enter Your Password')
        else:
            Found = False 
            for each_cell in range(3,passwordws.max_row +1 ):
                if a == passwordws['A'+str(each_cell)].value:
                    Found = True
                    break
            if Found == True:
                messagebox.showerror('Error', 'Username already exist. Please creae another username.')
                student_username_signup.delete(0,END)
                student_password_signup.delete(0,END)
            else: 
                max_row = passwordws.max_row+1

                passwordws.cell(row=max_row, column=1).value = a
                passwordws.cell(row=max_row, column=2).value = b
                passwordws.cell(row=max_row, column=3).value = c

                passwordwb.save('Passwords.xlsx')
                student_username_signup.delete(0,END)
                student_password_signup.delete(0,END)

            
                messagebox.showinfo('Success', 'Successfully created a new student account.')
                
        pass


    register_window.grid(row=0)
    register_window.grid_forget()

    
    ##################################################


    footer = Frame(root, width=1000, height=40, bg='#000000')
    footer.grid_propagate(False)
    footer.grid(row=2)

    Label(footer, text="@ 1980 Ambasing Intitute, Southern Mongolia", bg='#000000', fg='white' ).grid(row=0, column=0,padx=20,pady=10)

    Label(footer, text='ambasingfoundation@yahoo.com', bg='#000000', fg='white').grid(row=0, column=1,sticky=E,padx=500)




    def main_btn():
        main_button.configure(bg='#FFFBF5', fg='#000000')
        # reset
        login_button.configure(bg='#000000', fg='#FFFBF5')
        signin_button.configure(bg='#000000', fg='#FFFBF5')

        main_menu_body.grid(row=0)
        main_menu_body.grid_propagate(False)

        # coach.grid_forget()
        # student.grid_forget()
        register_window.grid_forget()
        main_body_frame.grid_forget()




    def login_btn():
        login_button.configure(bg='#FFFBF5', fg='#000000')
        # reset
        main_button.configure(bg='#000000', fg='#FFFBF5')
        signin_button.configure(bg='#000000', fg='#FFFBF5')

        main_body_frame.grid(row=0)
        main_body_frame.grid_propagate(False)
        coach.grid(row=0,column=0,padx=50,pady=75,sticky=NW)
        coach.grid_propagate(False)
        student.grid(row=0,column=0,padx=50,pady=75,sticky=NE)
        student.grid_propagate(False)


        main_menu_body.grid_forget() 
        register_window.grid_forget()


    global signin_btn
    def signin_btn():
        signin_button.configure(bg='#FFFBF5', fg='#000000')
        # reset
        main_button.configure(bg='#000000', fg='#FFFBF5')
        login_button.configure(bg='#000000', fg='#FFFBF5')

        register_window.grid(row=0)
        register_window.grid_propagate(False)

        main_menu_body.grid_forget()
        coach.grid_forget()
        student.grid_forget()
        main_body_frame.grid_forget()




def coach_menu():
    global coachmenu
    coachmenu = Toplevel()
    coachmenu.geometry('1000x700')
    coachmenu.title('Coach Menu')
    coachmenu.resizable(width=False,height=False)

    header = Frame(coachmenu, width=1000, height=50,bg='#525252')
    header.grid_propagate(False)
    header.grid(row=0)

    Button(header, text="⬅", font=('bold',15), bg='#FFFBF5', command=lambda:back()).grid(row=0, column=0, padx= 5, pady=5)
    Label(header, text='Health Database', font=('Arial Rounded MT Bold',25),fg='#FFFBF5',bg='#525252').grid(row=0, column=1, padx=311)
    Button(header,text="▢", font=('bold',15), bg='#FFFBF5', command=lambda:forget()).grid(row=0, column=2, padx=5 , pady=5, ipadx=3)

    def back():
        coachmenu.withdraw()
        root.deiconify()
        


    list = []
    def forget():

        counter = 1
        print(list)

        if len(list)%2 == 1: 
            print('even')
            list.append(counter)
            content.forget()
            print(list)

        elif len(list)%2 == 0:
            print('odd')
            list.append(counter)
            content.pack()
            print(list)






    body = Frame(coachmenu, width=1000, height=550,bg='#FFFBF5')
    body.grid_propagate(False)
    body.grid(row=1)

    content_left = Frame(body, width=300, height=550,bg='#FFFBF5')
    content_left.grid_propagate(False)
    content_left.grid(row=0,column=0)

    left_content = Frame(content_left,width=270, height=520,bg='#FFFBF5')
    left_content.grid_propagate(False)
    left_content.grid(padx=15,pady=15)

    Label(left_content, image=logo, width=270,height=270).grid(row=0,column=0,columnspan=2)

    Label(left_content, text='\n\n\nGive comment to your students.\n\n', font=('Arial Rounded MT Bold',13),bg='#FFFBF5',fg='black').grid(row=1,column=0,columnspan=2,sticky=N)
    Label(left_content,text='Search Last Name:', font=('Arial Rounded MT Bold',12),bg='#FFFBF5',fg='black').grid(row=2,column=0,sticky=E)

    comment_entry = Entry(left_content, width=10, font=('Arial Rounded MT Bold',12))
    comment_entry.grid(row=2,column=1,sticky=W)

    Button(left_content,text='search', font=('Arial Rounded MT Bold',13),fg='black',command=lambda:comment_button()).grid(row=3,column=0,columnspan=2,sticky=N,pady=20)

    def comment_button():
        global person
        person = comment_entry.get()
        Found = True
        for each_cell in range(5,database_ws.max_row+1):
            if comment_entry.get() == database_ws['Q'+str(each_cell)].value:

                cell_address = int(each_cell) - 4
                Found = False
                break
        if Found == False: 
            # messagebox.showinfo("Info", f'Data found in row: {cell_address}')
            
            

            def comment():
                    com = Toplevel()
                    com.geometry('300x200')
                    com.title('Comment Section')
                    com.resizable(width=False,height=False)
                    com.config(bg='white')


                    Label(com, text=f'\n\nComment to {person}:', font=('Arial Rounded MT Bold',13),bg='white').grid(row=0,sticky=N,padx=40)
                    phrase = Entry(com, width=20, font=('Arial Rounded MT Bold',13))
                    phrase.grid(row=1,sticky=N,padx=40,pady=10)
                    Button(com, text='comment',font=('Arial Rounded MT Bold',13),command=lambda:enter()).grid(row=2,sticky=N,pady=20)

                    def enter():
                        
                        database_ws['O'+str(each_cell)].value = phrase.get()

                        database_wb.save('Database.xlsx')

                        
                        com.destroy()

                        coachmenu.destroy()
                        coach_menu()


                        



                    com.mainloop()


            comment()        
            comment_entry.delete(0,END)

            

        else:
            messagebox.showerror("Info", f'There is no data for {person}')
            comment_entry.delete(0,END)



        pass

    content_top = Frame(body, width=700, height=400,bg='#FFFBF5')
    content_top.grid_propagate(False)
    content_top.grid(row=0,column=1,sticky=NW)

    top_content = Frame(content_top,width=670, height=370,bg='#414141')
    top_content.grid_propagate(False)
    top_content.grid(padx=15,pady=15)

    
    content = Frame(top_content, height=370, width=670, bg='#D7BBF5')
    content.pack_propagate(False)
    tv1 = ttk.Treeview(content, height=600)
    treescrolly = Scrollbar(content, orient="vertical", command=tv1.yview)
    treescrollx = Scrollbar(content, orient="horizontal", command=tv1.xview)
    tv1.configure(xscrollcommand = treescrollx.set, yscrollcommand=treescrolly.set)
    treescrollx.pack(side ="bottom",fill ="x")
    treescrolly.pack(side ="right",fill="y")  

    tv1['columns'] = ("Student No.", "Name", "Age", "Gender", "Contact", "Height", "Weight", "BMI","Nutrition(1st month)","Nutrition(2nd month)", "Nutrition(3rd month)", "Physique(1st month)", "Physique(2nd month)", "Physique(3rd month)", "Coach Comment")
    tv1.column("#0", width=1, minwidth=1)
    tv1.column("Student No.", anchor=CENTER, width=90)
    tv1.column("Name",  anchor=W, width=190)
    tv1.column("Age", anchor=CENTER, width=60)
    tv1.column("Gender", anchor=CENTER, width=60)
    tv1.column("Contact", anchor=CENTER, width=80)
    tv1.column("Height", anchor=CENTER, width=60)
    tv1.column("Weight", anchor=CENTER, width=60)
    tv1.column("BMI", anchor=CENTER, width=60)
    tv1.column("Nutrition(1st month)", anchor=CENTER, width=150)
    tv1.column("Nutrition(2nd month)", anchor=CENTER, width=150)
    tv1.column("Nutrition(3rd month)", anchor=CENTER, width=150)
    tv1.column("Physique(1st month)", anchor=CENTER, width=150)
    tv1.column("Physique(2nd month)", anchor=CENTER, width=150)
    tv1.column("Physique(3rd month)", anchor=CENTER, width=150)
    tv1.column("Coach Comment", anchor=W, width=120)

    tv1.heading("#0", text=" ", anchor=W)
    tv1.heading("Student No.", text="Student No.", anchor=CENTER)
    tv1.heading("Name", text="Name", anchor=CENTER)
    tv1.heading("Age", text="Age", anchor=CENTER)
    tv1.heading("Gender", text="Gender", anchor=CENTER)
    tv1.heading("Contact", text="Contact", anchor=CENTER)
    tv1.heading("Height", text="Height", anchor=CENTER)
    tv1.heading("Weight", text="Weight", anchor=CENTER)
    tv1.heading("BMI", text="BMI", anchor=CENTER)
    tv1.heading("Nutrition(1st month)", text="Nutrition(1st month)", anchor=CENTER)
    tv1.heading("Nutrition(2nd month)", text="Nutrition(2nd month)", anchor=CENTER)
    tv1.heading("Nutrition(3rd month)", text="Nutrition(3rd month)", anchor=CENTER)
    tv1.heading("Physique(1st month)", text="Physique(1st month)", anchor=CENTER)
    tv1.heading("Physique(2nd month)", text="Physique(2nd month)", anchor=CENTER)
    tv1.heading("Physique(3rd month)", text="Physique(3rd month)", anchor=CENTER)
    tv1.heading("Coach Comment", text="Coach Comment", anchor=CENTER)
    
    counter = 0
    for each_cell in range(5, (database_ws.max_row)+1):
        counter+=1
        tv1.insert(parent='', index="end",values=(database_ws['A'+str(each_cell)].value,database_ws['B'+str(each_cell)].value, database_ws['C'+str(each_cell)].value,database_ws['D'+str(each_cell)].value,database_ws['E'+str(each_cell)].value,database_ws['F'+str(each_cell)].value,database_ws['G'+str(each_cell)].value,database_ws['H'+str(each_cell)].value,database_ws['I'+str(each_cell)].value,database_ws['J'+str(each_cell)].value,database_ws['K'+str(each_cell)].value,database_ws['L'+str(each_cell)].value,database_ws['M'+str(each_cell)].value,database_ws['N'+str(each_cell)].value,database_ws['O'+str(each_cell)].value))
    tv1.pack()

    content.pack()
    content.forget()





    content_bottom = Frame(body, width=700, height=150,bg='#FFFBF5')
    content_bottom.grid_propagate(False)
    content_bottom.grid(row=0,column=1,sticky=SW)

    bottom_content = Frame(content_bottom,width=670, height=120,bg='#414141')
    bottom_content.grid_propagate(False)
    bottom_content.grid(padx=15,pady=15)

    search_ui = LabelFrame(bottom_content, width=670, height=120,text='Searching ...  ',fg='black',bg='#FFFBF5', font=('Arial Rounded MT Bold',16),borderwidth=2)
    search_ui.grid_propagate(False)
    search_ui.grid(row=0)




    Label(search_ui,text='     Search Last Name:\t', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=0,column=0,pady=10)
    search_entry = Entry(search_ui, width=15, font=('Arial Rounded MT Bold',14))
    search_entry.grid(row=0,column=1)
    Label(search_ui,text='Note: Search name if it exist.', font=('Arial Rounded MT Bold',10),bg='#FFFBF5').grid(row=1,column=0,columnspan=3,sticky=N,padx=180)
    Button(search_ui,text='search', font=('Arial Rounded MT Bold',12),bg='#FFFBF5',command=lambda:search()).grid(row=1,column=4,columnspan=2,sticky=E)
    


    
    edit_ui = LabelFrame(bottom_content, width=670, height=120,text='Editing ...  ',fg='black',bg='#FFFBF5', font=('Arial Rounded MT Bold',16),borderwidth=2)
    edit_ui.grid_propagate(False)
    edit_ui.grid(row=0)


    Label(edit_ui,text='     Search Last Name:\t', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=0,column=0,pady=10)
    edit_entry = Entry(edit_ui, width=15, font=('Arial Rounded MT Bold',14))
    edit_entry.grid(row=0,column=1)
    Label(edit_ui,text='Note: Search name if it exist.', font=('Arial Rounded MT Bold',10),bg='#FFFBF5').grid(row=1,column=0,columnspan=3,sticky=N,padx=180)
    Button(edit_ui,text='edit', font=('Arial Rounded MT Bold',12),bg='#FFFBF5',command=lambda:edit()).grid(row=1,column=4,columnspan=2,sticky=E,ipadx=13)
    



    
    delete_ui = LabelFrame(bottom_content, width=670, height=120,text='Deleting ...  ',fg='black',bg='#FFFBF5', font=('Arial Rounded MT Bold',16),borderwidth=2)
    delete_ui.grid_propagate(False)
    delete_ui.grid(row=0)


    Label(delete_ui,text='     Search Last Name:\t', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=0,column=0,pady=10)
    delete_entry = Entry(delete_ui, width=15, font=('Arial Rounded MT Bold',14))
    delete_entry.grid(row=0,column=1)
    Label(delete_ui,text='Note: Search name if it exist.', font=('Arial Rounded MT Bold',10),bg='#FFFBF5').grid(row=1,column=0,columnspan=3,sticky=N,padx=180)
    Button(delete_ui,text='delete', font=('Arial Rounded MT Bold',12),bg='#FFFBF5',command=lambda:delete()).grid(row=1,column=4,columnspan=2,sticky=E,ipadx=3)
    





    search_ui.grid_forget()
    edit_ui.grid_forget()
    delete_ui.grid_forget()

    footer = Frame(coachmenu, width=1000, height=100,bg='#525252')
    footer.grid_propagate(False)
    footer.grid(row=2)

    Button(footer, text="Add Student", font=('Arial Rounded MT Bold',13), bg='#FFFBF5', command=lambda:add_button()).grid(row= 0 , column= 0 ,ipadx=37, ipady=10, pady=20, padx=52) 
    Label(footer,text='\t', font=('Arial Rounded MT Bold',4), bg='#525252').grid(row=0,column=1)    
    Button(footer, text="Search Name", font=('Arial Rounded MT Bold',13), bg='#FFFBF5', command=lambda:search_button()).grid(row= 0 , column= 2 ,ipadx=36, ipady=10, pady=20, padx=12) 
    Button(footer, text="Edit Info", font=('Arial Rounded MT Bold',13), bg='#FFFBF5', command=lambda:edit_button()).grid(row= 0 , column= 3 ,ipadx=50, ipady=10, pady=20, padx=12) 
    Button(footer, text="Delete Info", font=('Arial Rounded MT Bold',13), bg='#FFFBF5', command=lambda:delete_button()).grid(row= 0 , column= 4 ,ipadx=40, ipady=10, pady=20, padx=12) 


    def add_button():
        coachmenu.withdraw()
        root.deiconify()
        signin_btn()
        pass


    def search_button():
        search_ui.grid_propagate(False)
        search_ui.grid(row=0)
        
        edit_ui.grid_forget()
        delete_ui.grid_forget()

    def edit_button():
        edit_ui.grid_propagate(False)
        edit_ui.grid(row=0)

        search_ui.grid_forget()
        delete_ui.grid_forget()
        


    def delete_button():
        delete_ui.grid_propagate(False)
        delete_ui.grid(row=0)

        search_ui.grid_forget()
        edit_ui.grid_forget()



        pass

    #############
    def search():
            
        a = search_entry.get()
        Found = True
        for each_cell in range(5,database_ws.max_row+1):
            if search_entry.get() == database_ws['Q'+str(each_cell)].value:

                cell_address = int(each_cell) - 4
                Found = False
                break
        if Found == False: 
            messagebox.showinfo("Info", f'Data found in row: {cell_address}')
            search_entry.delete(0,END)
        else:
            messagebox.showerror("Info", f'There is no data for {a}')
            search_entry.delete(0,END)

    def edit():

    


        # student_choices.withdraw()

        Found = True
        global each_cell1
        for each_cell1 in range(2,database_ws.max_row+1):
            if edit_entry.get() == database_ws['Q'+str(each_cell1)].value:
                global cell_address2
                cell_address2 = each_cell1
                Found = False
                break
        if Found == False: 
            # messagebox.showinfo("Info", f'Data found in row: {cell_address} ')   # There is no need for this anymore
            # messagebox.showinfo("Info","Success")
            

            


            student_update = Toplevel()
            student_update.geometry('800x700')
            student_update.resizable(width=FALSE, height=FALSE)
            # student_update.attributes('-topmost','true')
            student_update.title('Student update')

            # student_choices.withdraw()
            ############ Variables 09054030302


            update_number = database_ws['A'+str(each_cell1)].value
            update_username = database_ws['P'+str(each_cell1)].value

            update_fullname = database_ws['B'+str(each_cell1)].value
            update_age = database_ws['C'+str(each_cell1)].value
            update_gender = database_ws['D'+str(each_cell1)].value
            update_contact = database_ws['E'+str(each_cell1)].value

            update_height = database_ws['F'+str(each_cell1)].value
            update_weight = database_ws['G'+str(each_cell1)].value
            update_bmi = database_ws['H'+str(each_cell1)].value

            #Nutrition 
            update_nut_mon1 = database_ws['I'+str(each_cell1)].value
            update_nut_mon2 = database_ws['J'+str(each_cell1)].value
            update_nut_mon3 = database_ws['K'+str(each_cell1)].value

            #Physique
            update_phy_mon1 = database_ws['L'+str(each_cell1)].value
            update_phy_mon2 = database_ws['M'+str(each_cell1)].value
            update_phy_mon3 = database_ws['N'+str(each_cell1)].value




                    # database_ws['B'+str(each_cell1)].value = new_fullname.get()
                    # database_ws['C'+str(each_cell1)].value = new_age_var.get()
                    # database_ws['D'+str(each_cell1)].value = new_gender_var.get()
                    # database_ws['E'+str(each_cell1)].value = new_contact.get()


            fullname_str = StringVar()
            contact_str = StringVar()

            fullname_intvar = IntVar()
            age_intvar = IntVar()
            gender_intvar = IntVar()
            contact_intvar = IntVar()

            height_str = StringVar()
            weight_str = StringVar()
            bmi_str = StringVar()


            height_intvar = IntVar()
            weight_intvar = IntVar()
            bmi_intvar = IntVar()


            nut_mon1_intvar = IntVar()
            nut_mon2_intvar = IntVar()
            nut_mon3_intvar = IntVar()

            phy_mon1_intvar = IntVar()
            phy_mon2_intvar = IntVar()
            phy_mon3_intvar = IntVar()





            ############

            header = Frame(student_update, width=800, height=50)
            header.grid_propagate(FALSE)

            Button(header, text="⬅", font=('bold',15), command=lambda:back()).grid(row=0, column=0, padx=7)


            def back():
                student_update.withdraw()
                
                # student_choice()

            Button(header, text="Update", font=('bold',15), command=lambda:update_data_button()).grid(row=0,column=2, padx=7)

            def update_data_button():
                update_but()
                # student_update.withdraw()
                # student_choice()
                
                
                pass



            Label(header, text="[ Update Your Information ]", font=('bold',20)).grid(row=0, column=1,sticky=N,padx=157, pady=7)

            ###
            one = Frame(student_update, width=800, height=325)
            one.grid_propagate(FALSE)

            Label(one, text="        Basic Information", font=('bold',20),anchor=W).grid(row=0, column=0,sticky=W)
            Label(one, text="    ", font=('bold',25)).grid(row=0, column=1, columnspan=1,padx=20)
            # Button(one, text='Reset Entries', font=('bold',13),command=lambda:reset()).grid(row=0,column=2,columnspan=2,padx=10)

            def reset():
                pass

            content_1 = LabelFrame(one,width=400, height=140 )
            content_1.grid_propagate(FALSE)

            Label(content_1, text="Student Number:   ", font=('bold',15)).grid(row=0, column=0,sticky=E,pady=30, padx=10)
            Label(content_1,text=f"{update_number}", font=('bold',13)).grid(row=0, column=1)

            Label(content_1, text="Username:   ", font=('bold',15)).grid(row=1, column=0, sticky=E,padx=10)
            Label(content_1, text=f"{update_username}", font=('bold',13)).grid(row=1, column=1)


            ###
            content_2 = LabelFrame(one,width=400, height=140 )
            content_2.grid_propagate(FALSE)

            Label(content_2, text="Full Name:   ", font=('bold',15)).grid(row=0, column=0,sticky=E,pady=10, padx=10)
            Label(content_2,text=f"{update_fullname}", font=('bold',13)).grid(row=0, column=1)

            Label(content_2,text="Edit:   ", font=('bold',15)).grid(row=1, column=0, sticky=E, padx=10)
            #################### 
            new_fullname = Entry(content_2,textvariable=fullname_str, width=30, font=('bold',11))
            check_fullname = Checkbutton(content_2, text="   Same as before",font=("bold",10 ), variable=fullname_intvar, command=lambda:existing_fullname() )
            ####################
            new_fullname.grid(row=1, column=1)
            check_fullname.grid(row=2, column=1, pady=10)


            ###
            content_3 = LabelFrame(one,width=200, height=140 )
            content_3.grid_propagate(FALSE)


            Label(content_3, text="Age:   ", font=('bold',15)).grid(row=0, column=0,sticky=E,pady=10, padx=10)
            Label(content_3,text=f"{update_age}", font=('bold',13)).grid(row=0, column=1)

            Label(content_3,text="Edit:   ", font=('bold',15)).grid(row=1, column=0, sticky=E, padx=10)

            new_age_list = []
            print(new_age_list)
            for loop in range(1,101):
                new_age_list.append(loop)

            new_age_var = StringVar()

            ####################
            new_age = ttk.Combobox(content_3, textvariable=new_age_var, values=new_age_list,width=6, state='readonly',font=("bold",10 ) )
            check_age = Checkbutton(content_3, text="   Same as before",font=("bold",10 ), variable=age_intvar, command=lambda:existing_age() )
            ####################
            new_age.grid(row=1,column=1)
            check_age.grid(row=2, column=0,columnspan=2, pady=10, padx=10)




            ###
            content_4 = LabelFrame(one,width=200, height=140 )
            content_4.grid_propagate(FALSE)

            Label(content_4, text="Gender:   ", font=('bold',15)).grid(row=0, column=0,sticky=E,pady=10, padx=10)
            Label(content_4,text=f"{update_gender}", font=('bold',13)).grid(row=0, column=1)

            Label(content_4,text="Edit:   ", font=('bold',15)).grid(row=1, column=0, sticky=E, padx=10)


            new_gender_list = ["Male","Female"]
            new_gender_var = StringVar()

            ####################
            new_gender = ttk.Combobox(content_4, textvariable=new_gender_var, values=new_gender_list,width=6, state='readonly',font=("bold",10 ) )
            check_gender = Checkbutton(content_4, text="   Same as before",font=("bold",10 ), variable=gender_intvar, command=lambda:existing_gender() )
            ####################
            new_gender.grid(row=1,column=1)
            check_gender.grid(row=2, column=0,columnspan=2, pady=10, padx=10)



            ###
            content_5 = LabelFrame(one,width=400, height=140 )
            content_5.grid_propagate(FALSE)

            Label(content_5, text="Contact:   ", font=('bold',15)).grid(row=0, column=0,sticky=E,pady=10, padx=10)
            Label(content_5,text=f"{update_contact}", font=('bold',13)).grid(row=0, column=1)

            Label(content_5,text="Edit:   ", font=('bold',15)).grid(row=1, column=0, sticky=E, padx=10)
            #################### 
            new_contact = Entry(content_5,textvariable=contact_str, width=30, font=('bold',11))
            check_contact = Checkbutton(content_5, text="   Same as before",font=("bold",10 ), variable=contact_intvar, command=lambda:existing_contact() )
            ####################
            new_contact.grid(row=1, column=1)
            check_contact.grid(row=2, column=1, pady=10)

            ###
            content_1.grid(row=1,column=0)
            content_2.grid(row=2,column=0)
            content_3.grid(row=1,column=1)
            content_4.grid(row=1,column=2)
            content_5.grid(row=2,column=1, columnspan=2)


            ###
            two = Frame(student_update, width=800, height= 325)
            two.grid_propagate(FALSE)


            Label(two, text="         Health Information", font=('bold',20)).grid(row=0, columnspan=1,column=0,sticky=W)
            Label(two, text="    ", font=('bold',25)).grid(row=0, column=1,padx=40)
            # Button(two, text='Reset Entries', font=('bold',13),command=lambda:reset()).grid(row=0,column=1,stick=E,padx=43)

            
            cont1 = LabelFrame(two, width=300, height=280)
            cont1.grid_propagate(False)
            cont1.grid(row=1, column=0,sticky=W)


            # HEIGHT
            Label(cont1,text="Height:   ", font=('bold',13)).grid(row=0, column=0,sticky=E,pady=10, padx=10)
            Label(cont1,text=f"{update_height} cm", font=('bold',10)).grid(row=0, column=1)

            Label(cont1,text="Edit: ", font=('bold',13)).grid(row=1, column=0, sticky=E, padx=10,pady=8)
            new_height = Entry(cont1,textvariable=height_str, width=7, font=('bold',11))
            check_height = Checkbutton(cont1, text="   Same as before",font=("bold",9 ), variable=height_intvar, command=lambda:existing_height() )
            new_height.grid(row=1, column=1,sticky=W,padx=5)
            check_height.grid(row=1, column=2)

            # WEIGHT
            Label(cont1,text="Weight:   ", font=('bold',13)).grid(row=2, column=0,sticky=E,pady=10, padx=10)
            Label(cont1,text=f"{update_weight} kg", font=('bold',10)).grid(row=2, column=1)

            Label(cont1,text="Edit: ", font=('bold',13)).grid(row=3, column=0, sticky=E, padx=10,pady=8)
            new_weight = Entry(cont1,textvariable=weight_str, width=7, font=('bold',11))
            check_weight = Checkbutton(cont1, text="   Same as before",font=("bold",9 ), variable=weight_intvar, command=lambda:existing_weight() )
            new_weight.grid(row=3, column=1,sticky=W,padx=5)
            check_weight.grid(row=3, column=2)

            # BMI
            Label(cont1,text="BMI:   ", font=('bold',13)).grid(row=4, column=0,sticky=E,pady=10, padx=10)
            Label(cont1,text=f"{update_bmi}", font=('bold',10)).grid(row=4, column=1)

            Label(cont1,text="Edit: ", font=('bold',13)).grid(row=5, column=0, sticky=E, padx=10,pady=8)
            new_bmi = Entry(cont1,textvariable=bmi_str, width=7, font=('bold',11))
            check_bmi = Checkbutton(cont1, text="   Same as before",font=("bold",9 ), variable=bmi_intvar, command=lambda:existing_bmi() )
            new_bmi.grid(row=5, column=1,sticky=W,padx=5)
            check_bmi.grid(row=5, column=2)

    ###########################

            cont2= LabelFrame(two, width=500, height=140)
            cont2.grid_propagate(False)
            cont2.grid(row=1, column=1,sticky=NW)


            # Nutrition Month 1
            Label(cont2,text='Nutrition (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont2,text="Month 1: ", font=('bold',12)).grid(row=1, column=0,sticky=E,pady=10)
            Label(cont2,text=f"{update_nut_mon1}", font=('bold',10)).grid(row=1, column=1,sticky=N,ipady=13)

            Label(cont2,text="Edit: ", font=('bold',12)).grid(row=2, column=0, sticky=E)

            new_nutmon1_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_nutmon1_var = StringVar()

            new_nutmon1 = ttk.Combobox(cont2, textvariable=new_nutmon1_var, values=new_nutmon1_list,width=8, state='readonly',font=("bold",10 ) )
            check_nutmon1 = Checkbutton(cont2, text="   Same as before",font=("bold",10 ), variable=nut_mon1_intvar, command=lambda:existing_nut_month1() )
            new_nutmon1.grid(row=2,column=1,sticky=W,padx=10)
            check_nutmon1.grid(row=3, column=0,columnspan=2, pady=10, padx=10)

            # Nutrition month 2
            Label(cont2,text='Nutrition (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont2,text="Month 2: ", font=('bold',12)).grid(row=1, column=3,sticky=E,pady=10)
            Label(cont2,text=f"{update_nut_mon2}", font=('bold',10)).grid(row=1, column=4,sticky=N,ipady=13)

            Label(cont2,text="Edit: ", font=('bold',12)).grid(row=2, column=3, sticky=E)

            new_nutmon2_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_nutmon2_var = StringVar()

            new_nutmon2 = ttk.Combobox(cont2, textvariable=new_nutmon2_var, values=new_nutmon2_list,width=8, state='readonly',font=("bold",10 ) )
            check_nutmon2 = Checkbutton(cont2, text="   Same as before",font=("bold",10 ), variable=nut_mon2_intvar, command=lambda:existing_nut_month2() )
            new_nutmon2.grid(row=2,column=4,sticky=W,padx=10)
            check_nutmon2.grid(row=3, column=3,columnspan=2, pady=10, padx=10)

            # Nutrition Month 3
            Label(cont2,text='Nutrition (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont2,text="Month 3: ", font=('bold',12)).grid(row=1, column=5,sticky=E,pady=10)
            Label(cont2,text=f"{update_nut_mon3}", font=('bold',10)).grid(row=1, column=6,sticky=N,ipady=13)

            Label(cont2,text="Edit: ", font=('bold',12)).grid(row=2, column=5, sticky=E)

            new_nutmon3_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_nutmon3_var = StringVar()

            new_nutmon3 = ttk.Combobox(cont2, textvariable=new_nutmon3_var, values=new_nutmon3_list,width=8, state='readonly',font=("bold",10 ) )
            check_nutmon3 = Checkbutton(cont2, text="   Same as before",font=("bold",10 ), variable=nut_mon3_intvar, command=lambda:existing_nut_month3() )
            new_nutmon3.grid(row=2,column=6,sticky=W)
            check_nutmon3.grid(row=3, column=5,columnspan=2, pady=10, padx=10)



            cont3= LabelFrame(two, width=500, height=140)
            cont3.grid_propagate(False)
            cont3.grid(row=1, column=1,sticky=SW)

            # Physique Month 1
            Label(cont3,text='Physique (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont3,text="Month 1: ", font=('bold',12)).grid(row=1, column=0,sticky=E,pady=10)
            Label(cont3,text=f"{update_phy_mon1}", font=('bold',10)).grid(row=1, column=1,sticky=N,ipady=13)

            Label(cont3,text="Edit: ", font=('bold',12)).grid(row=2, column=0, sticky=E)

            new_phymon1_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_phymon1_var = StringVar()

            new_phymon1 = ttk.Combobox(cont3, textvariable=new_phymon1_var, values=new_phymon1_list,width=8, state='readonly',font=("bold",10 ) )
            check_phymon1 = Checkbutton(cont3, text="   Same as before",font=("bold",10 ), variable=phy_mon1_intvar, command=lambda:existing_phy_month1() )
            new_phymon1.grid(row=2,column=1,sticky=W,padx=10)
            check_phymon1.grid(row=3, column=0,columnspan=2, pady=10, padx=10)

            # Physique month 2
            Label(cont3,text='Physique (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont3,text="Month 2: ", font=('bold',12)).grid(row=1, column=3,sticky=E,pady=10)
            Label(cont3,text=f"{update_phy_mon2}", font=('bold',10)).grid(row=1, column=4,sticky=N,ipady=13)

            Label(cont3,text="Edit: ", font=('bold',12)).grid(row=2, column=3, sticky=E)

            new_phymon2_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_phymon2_var = StringVar()

            new_phymon2 = ttk.Combobox(cont3, textvariable=new_phymon2_var, values=new_phymon2_list,width=8, state='readonly',font=("bold",10 ) )
            check_phymon2 = Checkbutton(cont3, text="   Same as before",font=("bold",10 ), variable=phy_mon2_intvar, command=lambda:existing_phy_month2() )
            new_phymon2.grid(row=2,column=4,sticky=W,padx=10)
            check_phymon2.grid(row=3, column=3,columnspan=2, pady=10, padx=10)

            # Physique Month 3
            Label(cont3,text='Physique (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont3,text="Month 3: ", font=('bold',12)).grid(row=1, column=5,sticky=E,pady=10)
            Label(cont3,text=f"{update_phy_mon3}", font=('bold',10)).grid(row=1, column=6,sticky=N,ipady=13)

            Label(cont3,text="Edit: ", font=('bold',12)).grid(row=2, column=5, sticky=E)

            new_phymon3_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_phymon3_var = StringVar()

            new_phymon3 = ttk.Combobox(cont3, textvariable=new_phymon3_var, values=new_phymon3_list,width=8, state='readonly',font=("bold",10 ) )
            check_phymon3 = Checkbutton(cont3, text="   Same as before",font=("bold",10 ), variable=phy_mon3_intvar, command=lambda:existing_phy_month3() )
            new_phymon3.grid(row=2,column=6,sticky=W)
            check_phymon3.grid(row=3, column=5,columnspan=2, pady=10, padx=10)






            ###
            header.grid(row=1)
            one.grid(row=2)
            two.grid(row=3)

            
            ### def 

            def existing_fullname():
                if fullname_intvar.get() == 1:
                    old_fullname = database_ws['B'+str(each_cell1)].value
                    fullname_str.set(old_fullname)
                    new_fullname.config(state='disabled')
                
                
                elif fullname_intvar.get() == 0:
                    fullname_str.set("")
                    new_fullname.config(state='normal')

            def existing_age():
                if age_intvar.get() == 1:
                    old_reference = database_ws['C'+str(each_cell1)].value
                    new_age_var.set(old_reference)
                    new_age.config(state="disabled")
                
                
                elif age_intvar.get() == 0:
                    new_age_var.set("")
                    new_age.config(state='readonly')

            def existing_gender():
                if gender_intvar.get() == 1:
                    old_reference = database_ws['D'+str(each_cell1)].value
                    new_gender_var.set(old_reference)
                    new_gender.config(state="disabled")
                
                
                elif gender_intvar.get() == 0:
                    new_gender_var.set("")
                    new_gender.config(state='readonly')



            def existing_contact():
                if contact_intvar.get() == 1:
                    old_reference = database_ws['E'+str(each_cell1)].value
                    contact_str.set(old_reference)
                    new_contact.config(state="disabled")
                
                
                elif contact_intvar.get() == 0:
                    contact_str.set("")
                    new_contact.config(state='normal')

            def existing_height():
                if height_intvar.get() == 1:
                    old_height = database_ws['F'+str(each_cell1)].value
                    height_str.set(old_height)
                    new_height.config(state='disabled')
                
                
                elif height_intvar.get() == 0:
                    height_str.set("")
                    new_height.config(state='normal')
                
                pass


            def existing_weight():
                if weight_intvar.get() == 1:
                    old_weight = database_ws['G'+str(each_cell1)].value
                    weight_str.set(old_weight)
                    new_weight.config(state='disabled')
                
                
                elif weight_intvar.get() == 0:
                    weight_str.set("")
                    new_weight.config(state='normal')
                pass

            def existing_bmi():
                if bmi_intvar.get() == 1:
                    old_bmi = database_ws['H'+str(each_cell1)].value
                    bmi_str.set(old_bmi)
                    new_bmi.config(state='disabled')
                
                
                elif bmi_intvar.get() == 0:
                    bmi_str.set("")
                    new_bmi.config(state='normal')
                pass

            def existing_nut_month1():
                if nut_mon1_intvar.get() == 1:
                    old_reference = database_ws['I'+str(each_cell1)].value
                    new_nutmon1_var.set(old_reference)
                    new_nutmon1.config(state="disabled")
                
                
                elif nut_mon1_intvar.get() == 0:
                    new_nutmon1_var.set("")
                    new_nutmon1.config(state='readonly')

            def existing_nut_month2():
                if nut_mon2_intvar.get() == 1:
                    old_reference = database_ws['J'+str(each_cell1)].value
                    new_nutmon2_var.set(old_reference)
                    new_nutmon2.config(state="disabled")
                
                
                elif nut_mon2_intvar.get() == 0:
                    new_nutmon2_var.set("")
                    new_nutmon2.config(state='readonly')

            def existing_nut_month3():
                if nut_mon3_intvar.get() == 1:
                    old_reference = database_ws['K'+str(each_cell1)].value
                    new_nutmon3_var.set(old_reference)
                    new_nutmon3.config(state="disabled")
                
                
                elif nut_mon3_intvar.get() == 0:
                    new_nutmon3_var.set("")
                    new_nutmon3.config(state='readonly')


            def existing_phy_month1():
                if phy_mon1_intvar.get() == 1:
                    old_reference = database_ws['L'+str(each_cell1)].value
                    new_phymon1_var.set(old_reference)
                    new_phymon1.config(state="disabled")
                
                
                elif phy_mon1_intvar.get() == 0:
                    new_phymon1_var.set("")
                    new_phymon1.config(state='readonly')

            def existing_phy_month2():
                if phy_mon2_intvar.get() == 1:
                    old_reference = database_ws['M'+str(each_cell1)].value
                    new_phymon2_var.set(old_reference)
                    new_phymon2.config(state="disabled")
                
                
                elif phy_mon2_intvar.get() == 0:
                    new_phymon2_var.set("")
                    new_phymon2.config(state='readonly')

            def existing_phy_month3():
                if phy_mon3_intvar.get() == 1:
                    old_reference = database_ws['N'+str(each_cell1)].value
                    new_phymon3_var.set(old_reference)
                    new_phymon3.config(state="disabled")
                
                
                elif phy_mon3_intvar.get() == 0:
                    new_phymon3_var.set("")
                    new_phymon3.config(state='readonly')





            def update_but():

                a = new_fullname.get()
                b = new_age_var.get()
                c = new_gender_var.get()
                d = new_contact.get()

                e = new_height.get()
                f = new_weight.get()
                g = new_bmi.get()

                h = new_nutmon1_var.get()
                i = new_nutmon2_var.get()
                j = new_nutmon3_var.get()

                k = new_phymon1_var.get()
                l = new_phymon2_var.get()
                m = new_phymon3_var.get()






                if a == '' or b == '' or c == '' or d == '' or e == '' or f == '' or g == '' or h == '' or i == '' or j == '' or k == '' or l == '' or m == '':
                    messagebox.showerror("Error", "Please Fill all Entries.")

                else:
                    #Adding the data inside the Excel
                    database_ws['B'+str(each_cell1)].value = new_fullname.get()
                    database_ws['C'+str(each_cell1)].value = new_age_var.get()
                    database_ws['D'+str(each_cell1)].value = new_gender_var.get()
                    database_ws['E'+str(each_cell1)].value = new_contact.get()

                    database_ws['F'+str(each_cell1)].value = new_height.get()
                    database_ws['G'+str(each_cell1)].value = new_weight.get()
                    database_ws['H'+str(each_cell1)].value = new_bmi.get()

                    database_ws['I'+str(each_cell1)].value = new_nutmon1_var.get()
                    database_ws['J'+str(each_cell1)].value = new_nutmon2_var.get()
                    database_ws['K'+str(each_cell1)].value = new_nutmon3_var.get()

                    database_ws['L'+str(each_cell1)].value = new_phymon1_var.get()
                    database_ws['M'+str(each_cell1)].value = new_phymon2_var.get()
                    database_ws['N'+str(each_cell1)].value = new_phymon3_var.get()

                    database_ws['Q'+str(each_cell1)].value = edit_entry.get()




                    database_wb.save('Database.xlsx')
                    messagebox.showinfo("UPDATED","DATA HAS BEEN UPDATED")
                    student_update.withdraw()
                    coachmenu.destroy()
                    coach_menu()

                    
                    # student_choice()
                    









                    

            student_update.mainloop()

        else:
            messagebox.showerror("Info", f'There is no data for Username: {edit_entry.get()}')
        pass


    def delete():
        Found = False
        for each_cell in range(5, (database_ws.max_row) +1):
            if delete_entry.get() == database_ws["Q"+str(each_cell)].value:
                Found = True
                cell_address = each_cell
                break;
            else:
                Found = False
                
        if Found == False:
            messagebox.showerror("Info","Data don't exist. Unable to delete data.")
            delete_entry.delete(0,END) 

        if Found == True: 
            
            database_ws.delete_rows(cell_address)
            messagebox.showinfo("Info", "Data has been deleted.")
            delete_entry.delete(0,END)
            database_wb.save('Database.xlsx')

            coachmenu.destroy()
            coach_menu()
        

        







    coachmenu.mainloop()





def student_choice():
    student_choices = Toplevel()
    student_choices.geometry('600x600')
    student_choices.title('Student Menu')
    student_choices.resizable(width=False,height=False)
    

    frame1 = Frame(student_choices, width=600, height=70,)
    frame1.grid_propagate(FALSE)
    Label(frame1,text=' ', font=('Rockwell',5)).grid(row=0)
    Label(frame1, text=f'      Username: {global_username}\n', font=('Rockwell',17)).grid(row=1,column=0)
    frame1.grid(row=0)

    

    frame2 = LabelFrame(student_choices, text='[  Enroll  Data  ]', font=('Rockwell',17),height=150, width=600, borderwidth=3,labelanchor=NW)
    frame2.grid_propagate(FALSE)
    note = 'Note: You can only enroll once.'
    Label(frame2, font=('Rockwell',11), text=f'\n   You can add your basic information and health data here like the following: \n\t(Full Name, Age, Gender, Contact, Weight, Height, BMI)\n\n{note}\n').grid(row=0, column=0)
    Button(frame2, text='Enroll', font=('Rockwell',15), command=lambda:enroll()).grid(row=0,column=0, sticky=SE, ipadx=7, ipady=0)
    frame2.grid(row=1)


    frame3 = LabelFrame(student_choices, text='[  Update Information  ]', font=('Rockwell',17),height=150, width=600, borderwidth=3)
    frame3.grid_propagate(FALSE)
    Label(frame3, font=('Rockwell',11), text='\n You can update your personal and health information here like the following:\n\t(Full Name, Age, Gender, Contact, Weight, Height, BMI)\n\n\n').grid(row=0, column=0)
    Button(frame3, text='Update',command=lambda:student_update_window(), font=('Rockwell',15)).grid(row=0,column=0, sticky=SE, padx=0, ipadx=0, ipady=0)
    frame3.grid(row=2)


    frame5 = LabelFrame(student_choices, text='[  Show Personal Information  ]', font=('Rockwell',17),height=150, width=600, borderwidth=3)
    frame5.grid_propagate(FALSE)
    Label(frame5, font=('Rockwell',11), text='\n See your existing information about you.\t\t                                          \n \n \n\n').grid(row=0, column=0)
    Button(frame5, text='Show',command=lambda:show(), font=('Rockwell',15)).grid(row=0,column=0,sticky=SE, ipadx=10, ipady=0)
    frame5.grid(row=3)


    Button(student_choices, text='Logout', font=('Rockwell',20),command=lambda:back()).grid(row=4,pady=10,ipadx=5)


    def enroll():
        authorization = ''
        Found = False
        for each_cell in range(5, (database_ws.max_row) +1):
            if global_username == database_ws["P"+str(each_cell)].value:
                Found = True
                cell_address = each_cell
                break;
            else:
                Found = False
                
        if Found == False:
            # messagebox.showerror('Info','Enter your information once')
            # messagebox.showerror("Info","Data don't exist.")
            student_choices.withdraw()
            student_menu1()
            pass
            
        if Found == True: 
            messagebox.showinfo('Info','You can only enter your data once. \n But you can view/edit your data.')
            # messagebox.showinfo("Info", "Data Found")




    global student_update_window
    def student_update_window():


        # student_choices.withdraw()

        Found = True
        global each_cell1
        for each_cell1 in range(2,database_ws.max_row+1):
            if global_username == database_ws['P'+str(each_cell1)].value:
                global cell_address2
                cell_address2 = each_cell1
                Found = False
                break
        if Found == False: 

            student_update = Toplevel()
            student_update.geometry('800x700')
            student_update.resizable(width=FALSE, height=FALSE)
            # student_update.attributes('-topmost','true')
            student_update.title('Student update')

            student_choices.withdraw()
            ############ Variables 09054030302


            update_number = database_ws['A'+str(each_cell1)].value
            update_username = database_ws['P'+str(each_cell1)].value

            update_fullname = database_ws['B'+str(each_cell1)].value
            update_age = database_ws['C'+str(each_cell1)].value
            update_gender = database_ws['D'+str(each_cell1)].value
            update_contact = database_ws['E'+str(each_cell1)].value

            update_height = database_ws['F'+str(each_cell1)].value
            update_weight = database_ws['G'+str(each_cell1)].value
            update_bmi = database_ws['H'+str(each_cell1)].value

            #Nutrition 
            update_nut_mon1 = database_ws['I'+str(each_cell1)].value
            update_nut_mon2 = database_ws['J'+str(each_cell1)].value
            update_nut_mon3 = database_ws['K'+str(each_cell1)].value

            #Physique
            update_phy_mon1 = database_ws['L'+str(each_cell1)].value
            update_phy_mon2 = database_ws['M'+str(each_cell1)].value
            update_phy_mon3 = database_ws['N'+str(each_cell1)].value




                    # database_ws['B'+str(each_cell1)].value = new_fullname.get()
                    # database_ws['C'+str(each_cell1)].value = new_age_var.get()
                    # database_ws['D'+str(each_cell1)].value = new_gender_var.get()
                    # database_ws['E'+str(each_cell1)].value = new_contact.get()


            fullname_str = StringVar()
            contact_str = StringVar()

            fullname_intvar = IntVar()
            age_intvar = IntVar()
            gender_intvar = IntVar()
            contact_intvar = IntVar()

            height_str = StringVar()
            weight_str = StringVar()
            bmi_str = StringVar()


            height_intvar = IntVar()
            weight_intvar = IntVar()
            bmi_intvar = IntVar()


            nut_mon1_intvar = IntVar()
            nut_mon2_intvar = IntVar()
            nut_mon3_intvar = IntVar()

            phy_mon1_intvar = IntVar()
            phy_mon2_intvar = IntVar()
            phy_mon3_intvar = IntVar()





            ############

            header = Frame(student_update, width=800, height=50)
            header.grid_propagate(FALSE)

            Button(header, text="⬅", font=('bold',15), command=lambda:back()).grid(row=0, column=0, padx=7)


            def back():
                student_update.withdraw()
                
                student_choice()

            Button(header, text="Update", font=('bold',15), command=lambda:update_data_button()).grid(row=0,column=2, padx=7)

            def update_data_button():
                update_but()
                # student_update.withdraw()
                # student_choice()
                
                
                pass



            Label(header, text="[ Update Your Information ]", font=('bold',20)).grid(row=0, column=1,sticky=N,padx=157, pady=7)

            ###
            one = Frame(student_update, width=800, height=325)
            one.grid_propagate(FALSE)

            Label(one, text="        Basic Information", font=('bold',20),anchor=W).grid(row=0, column=0,sticky=W)
            Label(one, text="    ", font=('bold',25)).grid(row=0, column=1, columnspan=1,padx=20)
            # Button(one, text='Reset Entries', font=('bold',13),command=lambda:reset()).grid(row=0,column=2,columnspan=2,padx=10)

            def reset():
                pass

            content_1 = LabelFrame(one,width=400, height=140 )
            content_1.grid_propagate(FALSE)

            Label(content_1, text="Student Number:   ", font=('bold',15)).grid(row=0, column=0,sticky=E,pady=30, padx=10)
            Label(content_1,text=f"{update_number}", font=('bold',13)).grid(row=0, column=1)

            Label(content_1, text="Username:   ", font=('bold',15)).grid(row=1, column=0, sticky=E,padx=10)
            Label(content_1, text=f"{update_username}", font=('bold',13)).grid(row=1, column=1)


            ###
            content_2 = LabelFrame(one,width=400, height=140 )
            content_2.grid_propagate(FALSE)

            Label(content_2, text="Full Name:   ", font=('bold',15)).grid(row=0, column=0,sticky=E,pady=10, padx=10)
            Label(content_2,text=f"{update_fullname}", font=('bold',13)).grid(row=0, column=1)

            Label(content_2,text="Edit:   ", font=('bold',15)).grid(row=1, column=0, sticky=E, padx=10)
            #################### 
            new_fullname = Entry(content_2,textvariable=fullname_str, width=30, font=('bold',11))
            check_fullname = Checkbutton(content_2, text="   Same as before",font=("bold",10 ), variable=fullname_intvar, command=lambda:existing_fullname() )
            ####################
            new_fullname.grid(row=1, column=1)
            check_fullname.grid(row=2, column=1, pady=10)


            ###
            content_3 = LabelFrame(one,width=200, height=140 )
            content_3.grid_propagate(FALSE)


            Label(content_3, text="Age:   ", font=('bold',15)).grid(row=0, column=0,sticky=E,pady=10, padx=10)
            Label(content_3,text=f"{update_age}", font=('bold',13)).grid(row=0, column=1)

            Label(content_3,text="Edit:   ", font=('bold',15)).grid(row=1, column=0, sticky=E, padx=10)

            new_age_list = []
            print(new_age_list)
            for loop in range(1,101):
                new_age_list.append(loop)

            new_age_var = StringVar()

            ####################
            new_age = ttk.Combobox(content_3, textvariable=new_age_var, values=new_age_list,width=6, state='readonly',font=("bold",10 ) )
            check_age = Checkbutton(content_3, text="   Same as before",font=("bold",10 ), variable=age_intvar, command=lambda:existing_age() )
            ####################
            new_age.grid(row=1,column=1)
            check_age.grid(row=2, column=0,columnspan=2, pady=10, padx=10)




            ###
            content_4 = LabelFrame(one,width=200, height=140 )
            content_4.grid_propagate(FALSE)

            Label(content_4, text="Gender:   ", font=('bold',15)).grid(row=0, column=0,sticky=E,pady=10, padx=10)
            Label(content_4,text=f"{update_gender}", font=('bold',13)).grid(row=0, column=1)

            Label(content_4,text="Edit:   ", font=('bold',15)).grid(row=1, column=0, sticky=E, padx=10)


            new_gender_list = ["Male","Female"]
            new_gender_var = StringVar()

            ####################
            new_gender = ttk.Combobox(content_4, textvariable=new_gender_var, values=new_gender_list,width=6, state='readonly',font=("bold",10 ) )
            check_gender = Checkbutton(content_4, text="   Same as before",font=("bold",10 ), variable=gender_intvar, command=lambda:existing_gender() )
            ####################
            new_gender.grid(row=1,column=1)
            check_gender.grid(row=2, column=0,columnspan=2, pady=10, padx=10)



            ###
            content_5 = LabelFrame(one,width=400, height=140 )
            content_5.grid_propagate(FALSE)

            Label(content_5, text="Contact:   ", font=('bold',15)).grid(row=0, column=0,sticky=E,pady=10, padx=10)
            Label(content_5,text=f"{update_contact}", font=('bold',13)).grid(row=0, column=1)

            Label(content_5,text="Edit:   ", font=('bold',15)).grid(row=1, column=0, sticky=E, padx=10)
            #################### 
            new_contact = Entry(content_5,textvariable=contact_str, width=30, font=('bold',11))
            check_contact = Checkbutton(content_5, text="   Same as before",font=("bold",10 ), variable=contact_intvar, command=lambda:existing_contact() )
            ####################
            new_contact.grid(row=1, column=1)
            check_contact.grid(row=2, column=1, pady=10)

            ###
            content_1.grid(row=1,column=0)
            content_2.grid(row=2,column=0)
            content_3.grid(row=1,column=1)
            content_4.grid(row=1,column=2)
            content_5.grid(row=2,column=1, columnspan=2)


            ###
            two = Frame(student_update, width=800, height= 325)
            two.grid_propagate(FALSE)


            Label(two, text="         Health Information", font=('bold',20)).grid(row=0, columnspan=1,column=0,sticky=W)
            Label(two, text="    ", font=('bold',25)).grid(row=0, column=1,padx=40)
            # Button(two, text='Reset Entries', font=('bold',13),command=lambda:reset()).grid(row=0,column=1,stick=E,padx=43)

            
            cont1 = LabelFrame(two, width=300, height=280)
            cont1.grid_propagate(False)
            cont1.grid(row=1, column=0,sticky=W)


            # HEIGHT
            Label(cont1,text="Height:   ", font=('bold',13)).grid(row=0, column=0,sticky=E,pady=10, padx=10)
            Label(cont1,text=f"{update_height} cm", font=('bold',10)).grid(row=0, column=1)

            Label(cont1,text="Edit: ", font=('bold',13)).grid(row=1, column=0, sticky=E, padx=10,pady=8)
            new_height = Entry(cont1,textvariable=height_str, width=7, font=('bold',11))
            check_height = Checkbutton(cont1, text="   Same as before",font=("bold",9 ), variable=height_intvar, command=lambda:existing_height() )
            new_height.grid(row=1, column=1,sticky=W,padx=5)
            check_height.grid(row=1, column=2)

            # WEIGHT
            Label(cont1,text="Weight:   ", font=('bold',13)).grid(row=2, column=0,sticky=E,pady=10, padx=10)
            Label(cont1,text=f"{update_weight} kg", font=('bold',10)).grid(row=2, column=1)

            Label(cont1,text="Edit: ", font=('bold',13)).grid(row=3, column=0, sticky=E, padx=10,pady=8)
            new_weight = Entry(cont1,textvariable=weight_str, width=7, font=('bold',11))
            check_weight = Checkbutton(cont1, text="   Same as before",font=("bold",9 ), variable=weight_intvar, command=lambda:existing_weight() )
            new_weight.grid(row=3, column=1,sticky=W,padx=5)
            check_weight.grid(row=3, column=2)

            # BMI
            Label(cont1,text="BMI:   ", font=('bold',13)).grid(row=4, column=0,sticky=E,pady=10, padx=10)
            Label(cont1,text=f"{update_bmi}", font=('bold',10)).grid(row=4, column=1)

            Label(cont1,text="Edit: ", font=('bold',13)).grid(row=5, column=0, sticky=E, padx=10,pady=8)
            new_bmi = Entry(cont1,textvariable=bmi_str, width=7, font=('bold',11))
            check_bmi = Checkbutton(cont1, text="   Same as before",font=("bold",9 ), variable=bmi_intvar, command=lambda:existing_bmi() )
            new_bmi.grid(row=5, column=1,sticky=W,padx=5)
            check_bmi.grid(row=5, column=2)

    ###########################

            cont2= LabelFrame(two, width=500, height=140)
            cont2.grid_propagate(False)
            cont2.grid(row=1, column=1,sticky=NW)


            # Nutrition Month 1
            Label(cont2,text='Nutrition (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont2,text="Month 1: ", font=('bold',12)).grid(row=1, column=0,sticky=E,pady=10)
            Label(cont2,text=f"{update_nut_mon1}", font=('bold',10)).grid(row=1, column=1,sticky=N,ipady=13)

            Label(cont2,text="Edit: ", font=('bold',12)).grid(row=2, column=0, sticky=E)

            new_nutmon1_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_nutmon1_var = StringVar()

            new_nutmon1 = ttk.Combobox(cont2, textvariable=new_nutmon1_var, values=new_nutmon1_list,width=8, state='readonly',font=("bold",10 ) )
            check_nutmon1 = Checkbutton(cont2, text="   Same as before",font=("bold",10 ), variable=nut_mon1_intvar, command=lambda:existing_nut_month1() )
            new_nutmon1.grid(row=2,column=1,sticky=W,padx=10)
            check_nutmon1.grid(row=3, column=0,columnspan=2, pady=10, padx=10)

            # Nutrition month 2
            Label(cont2,text='Nutrition (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont2,text="Month 2: ", font=('bold',12)).grid(row=1, column=3,sticky=E,pady=10)
            Label(cont2,text=f"{update_nut_mon2}", font=('bold',10)).grid(row=1, column=4,sticky=N,ipady=13)

            Label(cont2,text="Edit: ", font=('bold',12)).grid(row=2, column=3, sticky=E)

            new_nutmon2_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_nutmon2_var = StringVar()

            new_nutmon2 = ttk.Combobox(cont2, textvariable=new_nutmon2_var, values=new_nutmon2_list,width=8, state='readonly',font=("bold",10 ) )
            check_nutmon2 = Checkbutton(cont2, text="   Same as before",font=("bold",10 ), variable=nut_mon2_intvar, command=lambda:existing_nut_month2() )
            new_nutmon2.grid(row=2,column=4,sticky=W,padx=10)
            check_nutmon2.grid(row=3, column=3,columnspan=2, pady=10, padx=10)

            # Nutrition Month 3
            Label(cont2,text='Nutrition (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont2,text="Month 3: ", font=('bold',12)).grid(row=1, column=5,sticky=E,pady=10)
            Label(cont2,text=f"{update_nut_mon3}", font=('bold',10)).grid(row=1, column=6,sticky=N,ipady=13)

            Label(cont2,text="Edit: ", font=('bold',12)).grid(row=2, column=5, sticky=E)

            new_nutmon3_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_nutmon3_var = StringVar()

            new_nutmon3 = ttk.Combobox(cont2, textvariable=new_nutmon3_var, values=new_nutmon3_list,width=8, state='readonly',font=("bold",10 ) )
            check_nutmon3 = Checkbutton(cont2, text="   Same as before",font=("bold",10 ), variable=nut_mon3_intvar, command=lambda:existing_nut_month3() )
            new_nutmon3.grid(row=2,column=6,sticky=W)
            check_nutmon3.grid(row=3, column=5,columnspan=2, pady=10, padx=10)



            cont3= LabelFrame(two, width=500, height=140)
            cont3.grid_propagate(False)
            cont3.grid(row=1, column=1,sticky=SW)

            # Physique Month 1
            Label(cont3,text='Physique (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont3,text="Month 1: ", font=('bold',12)).grid(row=1, column=0,sticky=E,pady=10)
            Label(cont3,text=f"{update_phy_mon1}", font=('bold',10)).grid(row=1, column=1,sticky=N,ipady=13)

            Label(cont3,text="Edit: ", font=('bold',12)).grid(row=2, column=0, sticky=E)

            new_phymon1_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_phymon1_var = StringVar()

            new_phymon1 = ttk.Combobox(cont3, textvariable=new_phymon1_var, values=new_phymon1_list,width=8, state='readonly',font=("bold",10 ) )
            check_phymon1 = Checkbutton(cont3, text="   Same as before",font=("bold",10 ), variable=phy_mon1_intvar, command=lambda:existing_phy_month1() )
            new_phymon1.grid(row=2,column=1,sticky=W,padx=10)
            check_phymon1.grid(row=3, column=0,columnspan=2, pady=10, padx=10)

            # Physique month 2
            Label(cont3,text='Physique (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont3,text="Month 2: ", font=('bold',12)).grid(row=1, column=3,sticky=E,pady=10)
            Label(cont3,text=f"{update_phy_mon2}", font=('bold',10)).grid(row=1, column=4,sticky=N,ipady=13)

            Label(cont3,text="Edit: ", font=('bold',12)).grid(row=2, column=3, sticky=E)

            new_phymon2_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_phymon2_var = StringVar()

            new_phymon2 = ttk.Combobox(cont3, textvariable=new_phymon2_var, values=new_phymon2_list,width=8, state='readonly',font=("bold",10 ) )
            check_phymon2 = Checkbutton(cont3, text="   Same as before",font=("bold",10 ), variable=phy_mon2_intvar, command=lambda:existing_phy_month2() )
            new_phymon2.grid(row=2,column=4,sticky=W,padx=10)
            check_phymon2.grid(row=3, column=3,columnspan=2, pady=10, padx=10)

            # Physique Month 3
            Label(cont3,text='Physique (Excellent, Good, Moderate, Bad)', font=('bold',13)).grid(row=0, column=0, columnspan=6)

            Label(cont3,text="Month 3: ", font=('bold',12)).grid(row=1, column=5,sticky=E,pady=10)
            Label(cont3,text=f"{update_phy_mon3}", font=('bold',10)).grid(row=1, column=6,sticky=N,ipady=13)

            Label(cont3,text="Edit: ", font=('bold',12)).grid(row=2, column=5, sticky=E)

            new_phymon3_list = ['Excellent', 'Good', 'Moderate', 'Bad']
            new_phymon3_var = StringVar()

            new_phymon3 = ttk.Combobox(cont3, textvariable=new_phymon3_var, values=new_phymon3_list,width=8, state='readonly',font=("bold",10 ) )
            check_phymon3 = Checkbutton(cont3, text="   Same as before",font=("bold",10 ), variable=phy_mon3_intvar, command=lambda:existing_phy_month3() )
            new_phymon3.grid(row=2,column=6,sticky=W)
            check_phymon3.grid(row=3, column=5,columnspan=2, pady=10, padx=10)






            ###
            header.grid(row=1)
            one.grid(row=2)
            two.grid(row=3)

            
            ### def 

            def existing_fullname():
                if fullname_intvar.get() == 1:
                    old_fullname = database_ws['B'+str(each_cell1)].value
                    fullname_str.set(old_fullname)
                    new_fullname.config(state='disabled')
                
                
                elif fullname_intvar.get() == 0:
                    fullname_str.set("")
                    new_fullname.config(state='normal')

            def existing_age():
                if age_intvar.get() == 1:
                    old_reference = database_ws['C'+str(each_cell1)].value
                    new_age_var.set(old_reference)
                    new_age.config(state="disabled")
                
                
                elif age_intvar.get() == 0:
                    new_age_var.set("")
                    new_age.config(state='readonly')

            def existing_gender():
                if gender_intvar.get() == 1:
                    old_reference = database_ws['D'+str(each_cell1)].value
                    new_gender_var.set(old_reference)
                    new_gender.config(state="disabled")
                
                
                elif gender_intvar.get() == 0:
                    new_gender_var.set("")
                    new_gender.config(state='readonly')



            def existing_contact():
                if contact_intvar.get() == 1:
                    old_reference = database_ws['E'+str(each_cell1)].value
                    contact_str.set(old_reference)
                    new_contact.config(state="disabled")
                
                
                elif contact_intvar.get() == 0:
                    contact_str.set("")
                    new_contact.config(state='normal')

            def existing_height():
                if height_intvar.get() == 1:
                    old_height = database_ws['F'+str(each_cell1)].value
                    height_str.set(old_height)
                    new_height.config(state='disabled')
                
                
                elif height_intvar.get() == 0:
                    height_str.set("")
                    new_height.config(state='normal')
                
                pass


            def existing_weight():
                if weight_intvar.get() == 1:
                    old_weight = database_ws['G'+str(each_cell1)].value
                    weight_str.set(old_weight)
                    new_weight.config(state='disabled')
                
                
                elif weight_intvar.get() == 0:
                    weight_str.set("")
                    new_weight.config(state='normal')
                pass

            def existing_bmi():
                if bmi_intvar.get() == 1:
                    old_bmi = database_ws['H'+str(each_cell1)].value
                    bmi_str.set(old_bmi)
                    new_bmi.config(state='disabled')
                
                
                elif bmi_intvar.get() == 0:
                    bmi_str.set("")
                    new_bmi.config(state='normal')
                pass

            def existing_nut_month1():
                if nut_mon1_intvar.get() == 1:
                    old_reference = database_ws['I'+str(each_cell1)].value
                    new_nutmon1_var.set(old_reference)
                    new_nutmon1.config(state="disabled")
                
                
                elif nut_mon1_intvar.get() == 0:
                    new_nutmon1_var.set("")
                    new_nutmon1.config(state='readonly')

            def existing_nut_month2():
                if nut_mon2_intvar.get() == 1:
                    old_reference = database_ws['J'+str(each_cell1)].value
                    new_nutmon2_var.set(old_reference)
                    new_nutmon2.config(state="disabled")
                
                
                elif nut_mon2_intvar.get() == 0:
                    new_nutmon2_var.set("")
                    new_nutmon2.config(state='readonly')

            def existing_nut_month3():
                if nut_mon3_intvar.get() == 1:
                    old_reference = database_ws['K'+str(each_cell1)].value
                    new_nutmon3_var.set(old_reference)
                    new_nutmon3.config(state="disabled")
                
                
                elif nut_mon3_intvar.get() == 0:
                    new_nutmon3_var.set("")
                    new_nutmon3.config(state='readonly')


            def existing_phy_month1():
                if phy_mon1_intvar.get() == 1:
                    old_reference = database_ws['L'+str(each_cell1)].value
                    new_phymon1_var.set(old_reference)
                    new_phymon1.config(state="disabled")
                
                
                elif phy_mon1_intvar.get() == 0:
                    new_phymon1_var.set("")
                    new_phymon1.config(state='readonly')

            def existing_phy_month2():
                if phy_mon2_intvar.get() == 1:
                    old_reference = database_ws['M'+str(each_cell1)].value
                    new_phymon2_var.set(old_reference)
                    new_phymon2.config(state="disabled")
                
                
                elif phy_mon2_intvar.get() == 0:
                    new_phymon2_var.set("")
                    new_phymon2.config(state='readonly')

            def existing_phy_month3():
                if phy_mon3_intvar.get() == 1:
                    old_reference = database_ws['N'+str(each_cell1)].value
                    new_phymon3_var.set(old_reference)
                    new_phymon3.config(state="disabled")
                
                
                elif phy_mon3_intvar.get() == 0:
                    new_phymon3_var.set("")
                    new_phymon3.config(state='readonly')





            def update_but():

                a = new_fullname.get()
                b = new_age_var.get()
                c = new_gender_var.get()
                d = new_contact.get()

                e = new_height.get()
                f = new_weight.get()
                g = new_bmi.get()

                h = new_nutmon1_var.get()
                i = new_nutmon2_var.get()
                j = new_nutmon3_var.get()

                k = new_phymon1_var.get()
                l = new_phymon2_var.get()
                m = new_phymon3_var.get()






                if a == '' or b == '' or c == '' or d == '' or e == '' or f == '' or g == '' or h == '' or i == '' or j == '' or k == '' or l == '' or m == '':
                    messagebox.showerror("Error", "Please Fill all Entries.")

                else:
                    #Adding the data inside the Excel
                    database_ws['B'+str(each_cell1)].value = new_fullname.get()
                    database_ws['C'+str(each_cell1)].value = new_age_var.get()
                    database_ws['D'+str(each_cell1)].value = new_gender_var.get()
                    database_ws['E'+str(each_cell1)].value = new_contact.get()

                    database_ws['F'+str(each_cell1)].value = new_height.get()
                    database_ws['G'+str(each_cell1)].value = new_weight.get()
                    database_ws['H'+str(each_cell1)].value = new_bmi.get()

                    database_ws['I'+str(each_cell1)].value = new_nutmon1_var.get()
                    database_ws['J'+str(each_cell1)].value = new_nutmon2_var.get()
                    database_ws['K'+str(each_cell1)].value = new_nutmon3_var.get()

                    database_ws['L'+str(each_cell1)].value = new_phymon1_var.get()
                    database_ws['M'+str(each_cell1)].value = new_phymon2_var.get()
                    database_ws['N'+str(each_cell1)].value = new_phymon3_var.get()





                    database_wb.save('Database.xlsx')
                    messagebox.showinfo("UPDATED","DATA HAS BEEN UPDATED")
                    student_update.withdraw()
                    student_choice()
                    









                    

            student_update.mainloop()

        else:
            messagebox.showerror("Info", f'There is no data for Username: {global_username}')



    def show():
        
        Found = False
        for each_cell in range(5, (database_ws.max_row) +1):
            if global_username == database_ws["P"+str(each_cell)].value:
                Found = True
                cell_address = each_cell
                break;
            else:
                Found = False
                
        if Found == False:
            messagebox.showerror("Info","Data don't exist.\nPlease Check if you have already enrolled.")
        if Found == True: 
            student_choices.withdraw()
            student_show()


    def back():
        student_choices.withdraw()
        root.deiconify()
        coachmenu.withdraw()
        


    student_choices.mainloop()

def student_menu1():
    global studentmenu1
    studentmenu1 = Toplevel()
    studentmenu1.geometry('500x600')
    studentmenu1.title('Student Menu')
    studentmenu1.resizable(width=False,height=False)

    header = Frame(studentmenu1, width=500, height= 100)
    header.grid_propagate(False)
    headertext = Frame(header, width=500, height=50)
    headertext.grid_propagate(False)
    Button(headertext, text="⬅", font=('bold',15), command=lambda:back()).grid(row=0, column=0, padx= 5, pady=5)
    Label(headertext,text='[ Basic Student Information ]', font=('bold',20)).grid(row=0, column= 1,padx=50 )
    def back():
        studentmenu1.withdraw()
        student_choice()
        pass
    headertext.grid(row=0)

    headertext2 = Frame(header, width=500, height=50)
    headertext2.grid_propagate(False)

    headertext2.grid(row=1)
    header.grid(row=0, column=0)

    content = Frame(studentmenu1, width=500, height= 450)
    content.grid_propagate(False)

    Label(content, text='Last Name : ', font=('bold',20)).grid(row=0, column=0, sticky=E,pady=8, ipady=5)
    Label(content, text='First Name : ', font=('bold',20)).grid(row=1, column=0, sticky=E,pady=8, ipady=5)
    Label(content, text='Middle Name : ', font=('bold',20)).grid(row=2, column=0, sticky=E,pady=8, ipady=5)
    Label(content, text='Age : ', font=('bold',20)).grid(row=3, column=0, sticky=E,pady=8, ipady=5)
    Label(content, text='Gender : ', font=('bold',20)).grid(row=4, column=0, sticky=E,pady=8, ipady=5)
    Label(content, text='Contact : ', font=('bold',20)).grid(row=5, column=0, sticky=E,pady=8, ipady=5)

    global last_name
    global first_name 
    global middle_name
    global age_var
    global gender_var
    global contact

    global last_name
    last_name = Entry(content, width=20, font=('bold',18) )
    first_name = Entry(content, width=20, font=('bold',18) )
    middle_name = Entry(content, width=20, font=('bold',18) )

    age_list = []
    print(age_list)
    for loop in range(1,101):
        age_list.append(loop)

    age_var = StringVar()
    age = ttk.Combobox(content, width=19, font=('bold',18),textvariable=age_var, values=age_list, state='readonly')

    gender_var = StringVar()
    female = Radiobutton(content,text='Female', font=('bold',18), value='Female', variable=gender_var)
    male = Radiobutton(content,text='Male', font=('bold',18), value='Male', variable=gender_var)
    contact = Entry(content, width=20, font=('bold',18) )

    last_name.grid(row=0, column=1, columnspan=2, sticky=W,pady=8, ipady=5)
    first_name.grid(row=1, column=1, columnspan=2, sticky=W,pady=8, ipady=5)
    middle_name.grid(row=2, column=1, columnspan=2, sticky=W,pady=8, ipady=5)
    age.grid(row=3, column=1, columnspan=2, sticky=W,pady=8, ipady=5)
    female.grid(row=4, column=1,pady=8, ipady=5)
    male.grid(row=4, column=2,pady=8, ipady=5)
    contact.grid(row=5, column=1, columnspan=2, sticky=W,pady=8, ipady=5)


    content.grid(row=1, column=0)

    footer = Frame(studentmenu1, width=500, height= 50)
    footer.grid_propagate(False)

    frame1 = Frame(footer, width=420, height=50)
    frame1.grid_propagate(False)
    frame1.grid(row=0,column=0)

    frame2 = Frame(footer, width=80, height=50)
    frame2.grid_propagate(False)
    Button(frame2, text='Next', font=('bold',15),command=lambda:nextt()).grid(padx=10, pady=5)
    def nextt(): 

        global student_number
        student_number = int(database_ws.cell(row=database_ws.max_row, column=1).value) + 1

       
        fn = first_name.get()
        ln = last_name.get()
        mn = middle_name.get()

        global fullname
        fullname = f'{ln}, {fn} {mn}'

        

        if last_name.get() == '' or first_name.get() == '' or middle_name.get() == '' or age_var.get() == '' or gender_var.get() == '' or  contact.get() == '':
            messagebox.showerror('Error','Please fill up all the entries.')
        else:

            Found = False
            for each_cell in range(5, database_ws.max_row + 1 ):
                if  fullname == database_ws['B'+str(each_cell)].value:
                    Found = True
                    break

          
            if Found == True: 
                messagebox.showinfo('Info','Name already exists.')
            else: 


                studentmenu1.withdraw()
                student_menu2()



    frame2.grid(row=0,column=1)

    footer.grid(row=2, column=0)


    studentmenu1.mainloop()

def student_menu2():
    studentmenu2 = Toplevel()
    studentmenu2.geometry('500x600')
    studentmenu2.title('Student Menu')
    studentmenu2.resizable(width=False,height=False)

    header = Frame(studentmenu2, width=500, height= 50,highlightthickness=1,highlightbackground='gray')
    header.grid_propagate(False)
    headertext = Frame(header, width=500, height=50)
    headertext.grid_propagate(False)
    Button(headertext, text="⬅", font=('bold',15), command=lambda:back()).grid(row=0, column=0, padx= 5, pady=5)
    Label(headertext,text='[ Student Health Information ]', font=('bold',20)).grid(row=0, column= 1,padx=50 )
    def back():

        studentmenu1.deiconify()
        studentmenu2.withdraw()

        pass
    headertext.grid(row=0)
    header.grid(row=0)

    content = Frame(studentmenu2, width=500, height= 500)
    content.grid_propagate(False)

    f1 = Frame(content, width=500,height=100, highlightthickness= 2, highlightbackground='gray')
    f1.grid_propagate(False)

    Label(f1, text='Height(cm) : ', font=('bold',15)).grid(row=0 ,column=0, padx=20, pady=15, sticky=N)
    Label(f1, text='Weight(kg) : ', font=('bold',15)).grid(row=0 ,column=1, padx=20, pady=15, sticky=N)
    Label(f1, text='BMI : ', font=('bold',15)).grid(row=0 ,column=2, padx=20, pady=15, sticky=N)

    height = Entry(f1, width=8, font=('bold',15))
    weight = Entry(f1, width=8, font=('bold',15))
    bmi = Entry(f1, width=8, font=('bold',15))

    height.grid(row=1, column=0, padx=38, pady=2, sticky=N)
    weight.grid(row=1, column=1, padx=38, pady=2, sticky=N)
    bmi.grid(row=1, column=2, padx=38, pady=2, sticky=N)

    f1.grid(row=0)

    f2 = Frame(content, width=500,height=200,highlightthickness=1,highlightbackground='gray')
    f2.grid_propagate(False)




    Label(f2, text=f'Self Assessment: Nutrition \n( Excellent , Good , Moderate , Bad )', font=('bold',15)).grid(row=0, column=0,columnspan=3, pady=7, sticky=N)
    Label(f2, text='*you can only fill-up the 1st month*', fg='red').grid(row=1, column=0, columnspan=3, pady=3)

    Label(f2, text='1st Month: ', font=('bold',15)).grid(row=2, column=0, padx=20, pady=10, sticky=N)
    Label(f2, text='2nd Month: ', font=('bold',15)).grid(row=2, column=1, padx=20, pady=10, sticky=N)
    Label(f2, text='3rd Month: ', font=('bold',15)).grid(row=2, column=2, padx=20, pady=10, sticky=N)

    global nut_var
    nutri_list = ["Excellent","Good","Moderate","Bad"]
    nut_var = StringVar()
    nutrition_month1 = ttk.Combobox(f2, width=10,textvariable=nut_var, values=nutri_list, font=('bold',15), state='readonly')
    nutrition_month2 = Entry(f2, width=11, font=('bold',15), state='readonly')
    nutrition_month3 = Entry(f2, width=11, font=('bold',15), state='readonly')

    nutrition_month1.grid(row=3, column=0, padx=17, pady=5, sticky=N)
    nutrition_month2.grid(row=3, column=1, padx=17, pady=5, sticky=N)
    nutrition_month3.grid(row=3, column=2, padx=17, pady=5, sticky=N)
    f2.grid(row=1)

    f3 = Frame(content, width=500,height=200, highlightthickness= 2, highlightbackground='gray')
    f3.grid_propagate(False)

    Label(f3, text=f'Self Assessment: Physique \n( Excellent , Good , Moderate , Bad )', font=('bold',15)).grid(row=0, column=0,columnspan=3, pady=7, sticky=N)
    Label(f3, text='*you can only fill-up the 1st month*', fg='red').grid(row=1, column=0, columnspan=3, pady=3)

    Label(f3, text='1st Month: ', font=('bold',15)).grid(row=2, column=0, padx=20, pady=10, sticky=N)
    Label(f3, text='2nd Month: ', font=('bold',15)).grid(row=2, column=1, padx=20, pady=10, sticky=N)
    Label(f3, text='3rd Month: ', font=('bold',15)).grid(row=2, column=2, padx=20, pady=10, sticky=N)

    ###########
    global phy_var
    phy_list = ["Excellent","Good","Moderate","Bad"]
    phy_var = StringVar()
    phy_month1 = ttk.Combobox(f3, width=10,textvariable=phy_var, values=phy_list, font=('bold',15), state='readonly')


 
    phy_month2 = Entry(f3, width=11, font=('bold',15), state='readonly')
    phy_month3 = Entry(f3, width=11, font=('bold',15), state='readonly')

    phy_month1.grid(row=3, column=0, padx=17, pady=5, sticky=N)
    phy_month2.grid(row=3, column=1, padx=17, pady=5, sticky=N)
    phy_month3.grid(row=3, column=2, padx=17, pady=5, sticky=N)


    f3.grid(row=2)
    content.grid(row=1)

    footer = Frame(studentmenu2, width=500, height= 50,highlightthickness=1,highlightbackground='gray')
    footer.grid_propagate(False)

    frame1 = Frame(footer, width=420, height=50)
    frame1.grid_propagate(False)
    frame1.grid(row=0,column=0)

    frame2 = Frame(footer, width=80, height=50)
    frame2.grid_propagate(False)
    Button(frame2, text='Finish', font=('bold',15), fg='black',command=lambda:finish()).grid(padx=10, pady=5)
    def finish():

        a = height.get()
        b = weight.get()
        c = bmi.get()
        d = nut_var.get()
        e = phy_var.get()


        if a == '' or b == '' or c == '' or d == '' or e == '':
            messagebox.showerror('Error','Please fill up all the entries.')
        else:


            maximum_row = database_ws.max_row + 1 

            database_ws.cell(row=maximum_row, column=1).value = student_number
            database_ws.cell(row=maximum_row, column=2).value = fullname
            database_ws.cell(row=maximum_row, column=3).value = age_var.get()
            database_ws.cell(row=maximum_row, column=4).value = gender_var.get()
            database_ws.cell(row=maximum_row, column=5).value = contact.get()


            database_ws.cell(row=maximum_row, column=6).value = height.get()
            database_ws.cell(row=maximum_row, column=7).value = weight.get()
            database_ws.cell(row=maximum_row, column=8).value = bmi.get()

            database_ws.cell(row=maximum_row, column=9).value = nut_var.get()
            database_ws.cell(row=maximum_row, column=12).value = phy_var.get()

            database_ws.cell(row=maximum_row, column=16).value = global_username
            database_ws.cell(row=maximum_row, column=17).value = last_name.get()
    

            messagebox.showinfo('Success', 'Your data has been successfully entered.')

            database_wb.save('Database.xlsx')

            studentmenu2.withdraw()
            student_choice()

    frame2.grid(row=0,column=1)
    footer.grid(row=2)
    studentmenu2.mainloop()





def student_show():
    stud_show = Toplevel()
    stud_show.geometry('1000x600')
    stud_show.title('Your Information')
    stud_show.resizable(width=False,height=False)


    #Searching for Rows
    Found = False
    for each_cell in range(5, (database_ws.max_row) +1):
        if global_username == database_ws["P"+str(each_cell)].value:
            Found = True
            cell_address = each_cell
            break;
        else:
            Found = False

    a = database_ws['A'+str(cell_address)].value
    b = database_ws['B'+str(cell_address)].value
    c = database_ws['C'+str(cell_address)].value
    d = database_ws['D'+str(cell_address)].value
    e = database_ws['E'+str(cell_address)].value
    f = database_ws['F'+str(cell_address)].value
    g = database_ws['G'+str(cell_address)].value
    h = database_ws['H'+str(cell_address)].value
    i = database_ws['I'+str(cell_address)].value
    j = database_ws['J'+str(cell_address)].value
    k = database_ws['K'+str(cell_address)].value
    l = database_ws['L'+str(cell_address)].value
    m = database_ws['M'+str(cell_address)].value
    n = database_ws['N'+str(cell_address)].value
    
    


    header = Frame(stud_show, width=1000, height=50,bg='#525252')
    header.grid_propagate(False)
    header.grid(row=0)

    Label(header, text='Your Information', font=('Arial Rounded MT Bold',20),bg='#525252',fg='#FFFBF5').grid(padx=400,pady=5)


    body = Frame(stud_show, width=1000, height=500,bg='#FFFBF5')
    body.grid_propagate(False)
    body.grid(row=1)

    content = Frame(body, width=900, height=400)
    content.grid_propagate(False)
    content.grid(row=0,padx=50,pady=50)

    con1 = Frame(content,width=450, height=400,bg='blue')
    con1.grid_propagate(False)
    con1.grid(row=0,column=0,sticky=W)

    aaa = LabelFrame(con1,width=450,height=400, text='Basic Information',labelanchor=N, font=('Arial Rounded MT Bold',20),bg='#FFFBF5')
    aaa.grid_propagate(False)
    aaa.grid()

    
    Label(aaa, text='  Student Number:  ', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=0,column=0,sticky=E,pady=7)
    Label(aaa, text='  Name:  ', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=1,column=0,sticky=E,pady=7)
    Label(aaa, text='  Age:  ', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=2,column=0,sticky=E,pady=7)
    Label(aaa, text='  Gender:  ', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=3,column=0,sticky=E,pady=7)
    Label(aaa, text='  Contact:  ', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=4,column=0,sticky=E,pady=7)
    Label(aaa, text='  Height:  ', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=5,column=0,sticky=E,pady=7)
    Label(aaa, text='  Weight:  ', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=6,column=0,sticky=E,pady=7)
    Label(aaa, text='  BMI:  ', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=7,column=0,sticky=E,pady=7)


    Label(aaa, text=f'{a}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=0,column=1,sticky=N,pady=7)
    Label(aaa, text=f'{b}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=1,column=1,sticky=N,pady=7)
    Label(aaa, text=f'{c}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=2,column=1,sticky=N,pady=7)
    Label(aaa, text=f'{d}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=3,column=1,sticky=N,pady=7)
    Label(aaa, text=f'{e}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=4,column=1,sticky=N,pady=7)
    Label(aaa, text=f'{f}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=5,column=1,sticky=N,pady=7)
    Label(aaa, text=f'{g}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=6,column=1,sticky=N,pady=7)
    Label(aaa, text=f'{h}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=7,column=1,sticky=N,pady=7)


    con2 = Frame(content,width=450, height=400,bg='yellow')
    con2.grid_propagate(False)
    con2.grid(row=0,column=1,sticky=E)

    bbb = LabelFrame(con2,width=450,height=400, text='Health Information',labelanchor=N, font=('Arial Rounded MT Bold',20),bg='#FFFBF5')
    bbb.grid_propagate(False)
    bbb.grid()

    Label(bbb, text='  [ Nutrition ] ', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=0,column=0,columnspan=2,sticky=W,pady=7)
    Label(bbb, text='\tMonth (1):\t', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=1,column=0,sticky=E,pady=7)
    Label(bbb, text='\tMonth (2):\t', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=2,column=0,sticky=E,pady=7)
    Label(bbb, text='\tMonth (3):\t', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=3,column=0,sticky=E,pady=7)

    Label(bbb, text='  [ Physique ] ', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=4,column=0,columnspan=2,sticky=W,pady=7)
    Label(bbb, text='\tMonth (1):\t', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=5,column=0,sticky=E,pady=7)
    Label(bbb, text='\tMonth (2):\t', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=6,column=0,sticky=E,pady=7)
    Label(bbb, text='\tMonth (3):\t', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=7,column=0,sticky=E,pady=7)

    Label(bbb, text=f'{i}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=1,column=1,sticky=W,pady=7)
    Label(bbb, text=f'{j}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=2,column=1,sticky=W,pady=7)
    Label(bbb, text=f'{k}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=3,column=1,sticky=W,pady=7)
    Label(bbb, text=f'{l}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=5,column=1,sticky=W,pady=7)
    Label(bbb, text=f'{m}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=6,column=1,sticky=W,pady=7)
    Label(bbb, text=f'{n}', font=('Arial Rounded MT Bold',15),bg='#FFFBF5').grid(row=7,column=1,sticky=W,pady=7)

    footer = Frame(stud_show, width=1000, height=50,bg='#525252')
    footer.grid_propagate(False)
    footer.grid(row=2)

    Button(footer, text='Go Back To Student Menu', command=lambda:back(), font=('Arial Rounded MT Bold',15)).grid(row=0, column=0,padx=380,pady=7)
    Button(footer, text='See Coach Comment', font=('Arial Rounded MT Bold',15),command=lambda:see()).grid(row=0,column=0,sticky=E,padx=50)

    def see():

        see_comment = database_ws['O'+str(cell_address)].value

        messagebox.showinfo('Comment',f'{see_comment}')
        pass

    def back():
        stud_show.withdraw()
        student_choice()
        

    stud_show.mainloop()




#CAll outs
main_window()
root.mainloop()